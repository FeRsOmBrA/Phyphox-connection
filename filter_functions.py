
from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd
import os 
from tkinter import filedialog
#########################################################################################


def create_excel(output_name):
    global output_path
    output_path = filedialog.askdirectory() 
    output_name = output_name + '.xlsx'
    output_path = output_path + '/' + output_name
    #create excel file
    global max_col, max_row, main_wb, main_ws
    main_wb = Workbook()
    main_ws = main_wb.active
    main_ws.title = 'Phyphox Sismic Exploration'
    main_ws['A1'] = 'Server Connection'
    main_ws['B1'] = 'System'
    main_ws['C1'] = 'Time'
    main_ws['D1'] = 'X'
    main_ws['E1'] = 'Y'
    main_ws['F1'] = 'Z'
    main_ws['G1'] = 'Start Time date'
    main_ws['H1'] = 'Start Time (s)'
    main_ws['I1'] = 'Stop Time date'
    main_ws['J1'] = 'Stop Time (s)'
    main_ws['K1'] = 'Real impact time (s)'
    max_col = main_ws.max_column
    max_row = main_ws.max_row
    main_wb.save(output_path)
    

def add_filter(path, system, sensibility, phone):
    accxiOS, accxAndroid = 'X (m/s^2)', 'Acceleration x (m/s^2)'
    accyiOS, accyAndroid = 'Y (m/s^2)', 'Acceleration y (m/s^2)'
    accziOS, acczAndroid = 'Z (m/s^2)', 'Acceleration z (m/s^2)'
    if system == 'Android':

        df_acc = pd.read_excel(path, sheet_name='Accelerometer')
        df_acc = df_acc[(df_acc[acczAndroid] < df_acc[acczAndroid].mean() - sensibility)]
        for i in range(0, len(df_acc)):
          main_ws['C' + str(max_row + phone)] =  df_acc['Time (s)'].iloc[i]
          main_ws['D' + str(max_row + phone)] =  df_acc[accxAndroid].iloc[i]
          main_ws['E' + str(max_row + phone)] =  df_acc[accyAndroid].iloc[i]
          main_ws['F' + str(max_row + phone)] =  df_acc[acczAndroid].iloc[i]
        meta_data = load_workbook(path)
        meta_data_time = meta_data['Metadata Time']
        server_ip = os.path.basename(path)
        #slice the extension
        server_ip = server_ip[:-5]
        
        main_ws[ 'A' + str(max_row + phone)] = server_ip
        main_ws['B' + str(max_row + phone)] = system 

        meta_data_start_date ='G' + str(max_row + phone)
        meta_data_start_time ='H' + str(max_row + phone)
        meta_data_stop_date ='I' + str(max_row + phone)
        meta_data_stop_time ='J' + str(max_row + phone)
        impact_time = 'C' + str(max_row + phone) 
        real_impact_time ='K' + str(max_row + phone)
        main_ws[meta_data_start_date ] = meta_data_time['D2'].value
        main_ws[meta_data_start_time ] =  F'=VALUE(MID({meta_data_start_date},18,6))'
        main_ws[meta_data_stop_date] = meta_data_time['D3'].value
        main_ws[meta_data_stop_time ] =  F'=VALUE(MID({meta_data_stop_date},18,6))'
        main_ws[real_impact_time] =  f'={impact_time}+ {meta_data_start_time}'

        
    elif system == 'iOS':

        df_acc = pd.read_excel(path, sheet_name='Accelerometer')
        df_acc = df_acc[(df_acc[accziOS] < df_acc[accziOS].mean() - sensibility)]
        for i in range(0, len(df_acc)):
          main_ws['C' + str(max_row + phone)] =  df_acc['Time (s)'].iloc[i]
          main_ws['D' + str(max_row + phone)] =  df_acc[accxiOS].iloc[i]
          main_ws['E' + str(max_row + phone)] =  df_acc[accyiOS].iloc[i]
          main_ws['F' + str(max_row + phone)] =  df_acc[accziOS].iloc[i]
        meta_data = load_workbook(path)
        meta_data_time = meta_data['Metadata Time']
        server_ip = os.path.basename(path)
        #slice the extension
        server_ip = server_ip[:-5]
        
        main_ws[ 'A' + str(max_row + phone)] = server_ip
        main_ws['B' + str(max_row + phone)] = system
        meta_data_start_date ='G' + str(max_row + phone)
        meta_data_start_time ='H' + str(max_row + phone)
        meta_data_stop_date ='I' + str(max_row + phone)
        meta_data_stop_time ='J' + str(max_row + phone)
        impact_time = 'C' + str(max_row + phone) 
        real_impact_time ='K' + str(max_row + phone)
        main_ws[meta_data_start_date ] = meta_data_time['D2'].value
        main_ws[meta_data_start_time ] =  F'=VALUE(MID({meta_data_start_date},18,6))'
        main_ws[meta_data_stop_date] = meta_data_time['D3'].value
        main_ws[meta_data_stop_time ] =  F'=VALUE(MID({meta_data_stop_date},18,6))'
        main_ws[real_impact_time] =  f'={impact_time}+ {meta_data_start_time}'  
    
    else : 
        print('ERORR')
    main_wb.save(output_path) 

def wave_arrival_time():

    main_ws['L1'] = 'Wave arrival time (s)'
    for i in range (main_ws.max_row -2):
            main_ws['L2'] = 0 
            k_b = 'K' + str(i+2)
            k_a = 'K' + str(i+3)
            main_ws['L' + str(i+3)] = f'={k_b}-{k_a}'
    main_wb.save(output_path)
    os.startfile(output_path)

create_excel('fer')
add_filter( path= 'C:/Users/PERSONAL/Desktop/Phyphox/172.20.10.1.xlsx' , system= 'iOS' , sensibility= 0.5, phone= 1)