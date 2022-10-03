

import pandas as pd
import tkinter as tk
from tkinter import  ttk
from tkinter import messagebox
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)
import pandas as pd
import os 
from tkinter import filedialog
from ttkthemes import ThemedTk
def main():

    def select_file():

        path = filedialog.askopenfilename(initialdir=".", title="Selecciona el archivo, debe ser xls o xlsx", filetypes=(
            ("Excel files", "*.xlsx"), ("Excel files", "*.xls"), ("All files", "*.*")))
        # extract only the name of the file
        file_path.set(path)
        file_name = path.split('/')[-1]
        if file_path.get() == '':

            label_file_path.config(text='No se ha seleccionado ningún archivo')
        else:

            label_file_path.config(
                text='El archivo seleccionado es' + ' ' + file_name)

    def read_file():
        if file_path.get() == "":
            messagebox.showerror('Error', 'Selecciona un archivo')
        elif file_path.get() != "" and file_path.get().endswith('.xls') or file_path.get().endswith('.xlsx'):
            messagebox.showinfo(
                'Exito', 'El archivo tiene la extensión adecuada, imprimiendo datos...')
            df = pd.read_excel(file_path.get())
            label_file_path.config(text=df)
        elif file_path.get().endswith('.py'):
            messagebox.showerror(
                'Error', 'Emm, esto es un script de python, no un archivo de excel')

        elif file_path.get().endswith('.csv'):
            messagebox.showerror(
                'Error', 'Emm, esto es un archivo csv, no un archivo de excel')

        elif file_path.get().endswith('.txt'):
            messagebox.showerror(
                'Error', 'Emm, esto es un archivo txt, no un archivo de excel')
        elif file_path.get().endswith('.docx'):
            messagebox.showerror(
                'Error', 'Emm, esto es un archivo de word, no un archivo de excel')
        elif file_path.get().endswith('.pdf'):
            messagebox.showerror(
                'Error', 'Emm, esto es un archivo pdf, no un archivo de excel')
        elif file_path.get().endswith('.html'):
            messagebox.showwerror(
                'Error', 'Emm, esto es un archivo html, no un archivo de excel')
        elif file_path.get().endswith('.xml'):
            messagebox.showerror(
                'Error', 'Emm, esto es un archivo xml, no un archivo de excel')
        elif file_path.get().endswith('.png') or file_path.get().endswith('.jpg') or file_path.get().endswith('.jpeg'):
            messagebox.showerror(
                'Error', 'Emm, esto es una imagen, no un archivo de excel')
        elif file_path.get().endswith(''):
            messagebox.showerror(
                'Error', 'Emm, este archivo nisiquiera tiene extensión')
        else:
            messagebox.showerror(
                'Error', f'El archivo tiene la extension .{file_path.get().split(".")[-1]} por favor, selecciona un archivo con la extensión xlsx o xls')

    def filter_data():
           global df
           try:
                 accziOS, acczAndroid = 'Z (m/s^2)', 'Acceleration z (m/s^2)'
                 if entry.get() == "android" or entry.get() == "Android":
                     df = pd.read_excel(file_path.get())
                     df = df[(df[acczAndroid] < df[acczAndroid].mean() - sensibility.get())]
                     # show only the time and the acceleration
                     if df.empty == True:
                         label_file_path.config(text='No hay anomalias en los datos')
                     elif sensibility.get() == 0 and df.empty == True:
                        label_file_path.config(text='No hay anomalias en los datos')
                     elif sensibility.get() == 0:
                        label_file_path.config(text='La sensibilidad no puede ser 0')

                     else:
                            label_file_path.config(text=df)


                 elif entry.get() == "ios" or entry.get() == "iOS":
                     df = pd.read_excel(file_path.get())
         
                     df = df[(df[accziOS] < df[accziOS].mean() - sensibility.get())]
                     # show only the time and the acceleration
                     df = df[['Time (s)', accziOS]]
                     if df.empty == True:
                         label_file_path.config(text='No hay anomalias en los datos')
                     elif sensibility.get() == 0 and df.empty == True:
                        label_file_path.config(text='No hay anomalias en los datos')
                     elif sensibility.get() == 0:
                        label_file_path.config(text='La sensibilidad no puede ser 0')

                     else:
                            label_file_path.config(text=df)
         
                 elif file_path.get() == "":
                     messagebox.showerror('Error', 'Selecciona un archivo')
                 else:
                     messagebox.showerror(
                         'Error', 'Debes seleccionar el sistema operativo del dispositivo')
           except:
                 messagebox.showerror('Error', 'Revisa la selección del sistema operativo')

    def Show_add_filter():
            phyphox_sismic_exploration_main.destroy()
            phyphox_sismic_exploration_filter = ThemedTk(theme='arc')
            phyphox_sismic_exploration_filter.title('Phyphox Sismic Exploration Filter')
            phyphox_sismic_exploration_filter.state('zoomed')
            phyphox_sismic_exploration_filter.attributes('-fullscreen', True)
            phyphox_sismic_exploration_filter.config(padx=10, pady=10)
            phyphox_sismic_exploration_filter.resizable(0, 0)
            def create_excel(output_name):
                global output_path
                output_path = filedialog.askdirectory() 
                output_name = output_name + '.xlsx'
                output_path = output_path + '/' + output_name
                #create excel file
                global max_col, max_row, main_wb, main_ws
                main_wb = Workbook()
                main_ws = main_wb.active
                main_ws.title = 'Phyphox Sismic Exploration'
                main_ws['A1'] = 'Server Connection'
                main_ws['B1'] = 'System'
                main_ws['C1'] = 'Time'
                main_ws['D1'] = 'X'
                main_ws['E1'] = 'Y'
                main_ws['F1'] = 'Z'
                main_ws['G1'] = 'Start Time date'
                main_ws['H1'] = 'Start Time (s)'
                main_ws['I1'] = 'Stop Time date'
                main_ws['J1'] = 'Stop Time (s)'
                main_ws['K1'] = 'Real impact time (s)'
                max_col = main_ws.max_column
                max_row = main_ws.max_row
                main_wb.save(output_path)
                
            
            def add_filter(path, system, sensibility, phone):

                accxiOS, accxAndroid = 'X (m/s^2)', 'Acceleration x (m/s^2)'
                accyiOS, accyAndroid = 'Y (m/s^2)', 'Acceleration y (m/s^2)'
                accziOS, acczAndroid = 'Z (m/s^2)', 'Acceleration z (m/s^2)'
                if system == 'Android':
                    
                    df_acc = pd.read_excel(path, sheet_name='Accelerometer')
                    df_acc = df_acc[(df_acc[acczAndroid] < df_acc[acczAndroid].mean() - sensibility)]
                    if df_acc.empty == True:
                            for i in range(0, 1):
                                 main_ws['C' + str(max_row + phone)] =  'No hay anomalias en los datos'
                                 main_ws['D' + str(max_row + phone)] =  'No hay anomalias en los datos'
                                 main_ws['E' + str(max_row + phone)] =  'No hay anomalias en los datos'
                                 main_ws['F' + str(max_row + phone)] =  'No hay anomalias en los datos'

        
                            meta_data = load_workbook(path)
                            meta_data_time = meta_data['Metadata Time']
                            server_ip = os.path.basename(path)
                            #slice the extension
                            server_ip = server_ip[:-5]
                            main_ws[ 'A' + str(max_row + phone)] = server_ip
                            main_ws['B' + str(max_row + phone)] = system
                            meta_data_start_date ='G' + str(max_row + phone)
                            meta_data_start_time ='H' + str(max_row + phone)
                            meta_data_stop_date ='I' + str(max_row + phone)
                            meta_data_stop_time ='J' + str(max_row + phone)
                            impact_time = 'C' + str(max_row + phone) 
                            real_impact_time ='K' + str(max_row + phone)
                            main_ws[meta_data_start_date ] = meta_data_time['D2'].value
                            main_ws[meta_data_start_time ] =  F'=VALUE(MID({meta_data_start_date},18,6))'
                            main_ws[meta_data_stop_date] = meta_data_time['D3'].value
                            main_ws[meta_data_stop_time ] =  F'=VALUE(MID({meta_data_stop_date},18,6))'
                            main_ws[real_impact_time] =  'No hay anomalias en los datos'
                    else:
                            for i in range(0, len(df_acc)):
                                 main_ws['C' + str(max_row + phone)] =  df_acc['Time (s)'].iloc[i]
                                 main_ws['D' + str(max_row + phone)] =  df_acc[accxAndroid].iloc[i]
                                 main_ws['E' + str(max_row + phone)] =  df_acc[accyAndroid].iloc[i]
                                 main_ws['F' + str(max_row + phone)] =  df_acc[acczAndroid].iloc[i]

                    print('done for android')
            
                    
                elif system == 'iOS':
            
                    df_acc = pd.read_excel(path, sheet_name='Accelerometer')
                    df_acc = df_acc[(df_acc[accziOS] < df_acc[accziOS].mean() - sensibility)]
                    if df_acc.empty == True:
                            for i in range(0, 1):
                                 main_ws['C' + str(max_row + phone)] =  'No hay anomalias en los datos'
                                 main_ws['D' + str(max_row + phone)] =  'No hay anomalias en los datos'
                                 main_ws['E' + str(max_row + phone)] =  'No hay anomalias en los datos'
                                 main_ws['F' + str(max_row + phone)] =  'No hay anomalias en los datos'

        
                            meta_data = load_workbook(path)
                            meta_data_time = meta_data['Metadata Time']
                            server_ip = os.path.basename(path)
                            #slice the extension
                            server_ip = server_ip[:-5]
                            main_ws[ 'A' + str(max_row + phone)] = server_ip
                            main_ws['B' + str(max_row + phone)] = system
                            meta_data_start_date ='G' + str(max_row + phone)
                            meta_data_start_time ='H' + str(max_row + phone)
                            meta_data_stop_date ='I' + str(max_row + phone)
                            meta_data_stop_time ='J' + str(max_row + phone)
                            impact_time = 'C' + str(max_row + phone) 
                            real_impact_time ='K' + str(max_row + phone)
                            main_ws[meta_data_start_date ] = meta_data_time['D2'].value
                            main_ws[meta_data_start_time ] =  F'=VALUE(MID({meta_data_start_date},18,6))'
                            main_ws[meta_data_stop_date] = meta_data_time['D3'].value
                            main_ws[meta_data_stop_time ] =  F'=VALUE(MID({meta_data_stop_date},18,6))'
                            main_ws[real_impact_time] =  'No hay anomalias en los datos'
                    else:
                            for i in range(0, len(df_acc)):
                                 main_ws['C' + str(max_row + phone)] =  df_acc['Time (s)'].iloc[i]
                                 main_ws['D' + str(max_row + phone)] =  df_acc[accxiOS].iloc[i]
                                 main_ws['E' + str(max_row + phone)] =  df_acc[accyiOS].iloc[i]
                                 main_ws['F' + str(max_row + phone)] =  df_acc[accziOS].iloc[i]

        
                            meta_data = load_workbook(path)
                            meta_data_time = meta_data['Metadata Time']
                            server_ip = os.path.basename(path)
                            #slice the extension
                            server_ip = server_ip[:-5]
                            main_ws[ 'A' + str(max_row + phone)] = server_ip
                            main_ws['B' + str(max_row + phone)] = system
                            meta_data_start_date ='G' + str(max_row + phone)
                            meta_data_start_time ='H' + str(max_row + phone)
                            meta_data_stop_date ='I' + str(max_row + phone)
                            meta_data_stop_time ='J' + str(max_row + phone)
                            impact_time = 'C' + str(max_row + phone) 
                            real_impact_time ='K' + str(max_row + phone)
                            main_ws[meta_data_start_date ] = meta_data_time['D2'].value
                            main_ws[meta_data_start_time ] =  F'=VALUE(MID({meta_data_start_date},18,6))'
                            main_ws[meta_data_stop_date] = meta_data_time['D3'].value
                            main_ws[meta_data_stop_time ] =  F'=VALUE(MID({meta_data_stop_date},18,6))'
                            main_ws[real_impact_time] =  f'={impact_time}+ {meta_data_start_time}' 
                            
                    print('done for ios') 
                else: 
                    messagebox.showerror('Error', 'Selecciona un sistema operativo')
             
            def save_file():
                try:
                
                   main_wb.save(output_path)
                   print('File saved')
                except:
                    messagebox.showerror('Error', 'Primero debes crear el archivo de salida') 
    

            def Start_filter():

                filters_frame_header = ttk.Frame(phyphox_sismic_exploration_filter)
                filters_frame_header.pack()
              
                init_label = tk.Label(filters_frame_header, text='Ahora debes seleccionar los archivos de excel de cada telefono y especificar su sistema operativo')
                init_label.grid(row=3, column=0)                
                init_label.config(font=('Cascadia Code', 9),fg = 'black')
                filters_frame = tk.Frame(phyphox_sismic_exploration_filter)
                filters_frame.pack( )    
                filters_frame.config(  padx=10, pady=10)
                filters1 = tk.Label(filters_frame, text=f'Archivo {1}')
                filters1.grid(row=0, column=1)
                filters1.config(font=('Cascadia Code', 9),fg = 'black')
                filters2 = tk.Label(filters_frame, text=f'Archivo {2}')
                filters2.grid(row=1, column=1)
                filters2.config(font=('Cascadia Code', 9),fg = 'black')
                filters3 = tk.Label(filters_frame, text=f'Archivo {3}')
                filters3.grid(row=2, column=1)
                filters3.config(font=('Cascadia Code', 9),fg = 'black')
                filters4 = tk.Label(filters_frame, text=f'Archivo {4}')
                filters4.grid(row=3, column=1)
                filters4.config(font=('Cascadia Code', 9),fg = 'black')
                filters5 = tk.Label(filters_frame, text=f'Archivo {5}')
                filters5.grid(row=4, column=1)
                filters5.config(font=('Cascadia Code', 9),fg = 'black')
                filters6 = tk.Label(filters_frame, text=f'Archivo {6}')
                filters6.grid(row=5, column=1)
                filters6.config(font=('Cascadia Code', 9),fg = 'black')
                filters7 = tk.Label(filters_frame, text=f'Archivo {7}')
                filters7.grid(row=6, column=1)
                filters7.config(font=('Cascadia Code', 9),fg = 'black')
                filters8 = tk.Label(filters_frame, text=f'Archivo {8}')
                filters8.grid(row=7, column=1)
                filters8.config(font=('Cascadia Code', 9),fg = 'black')
                filters9 = tk.Label(filters_frame, text=f'Archivo {9}')
                filters9.grid(row=8, column=1)
                filters9.config(font=('Cascadia Code', 9),fg = 'black')
                filters10 = tk.Label(filters_frame, text=f'Archivo {10}')
                filters10.grid(row=9, column=1)
                filters10.config(font=('Cascadia Code', 9),fg = 'black')


                def select_path1():
                    global file_path1
                    file_path1 = filedialog.askopenfilename(
                        initialdir='.', title="Select A File", filetypes=(("Excel Files", "*.xlsx"), ("Excel Files", "*.xls")))
                    file_name1 = os.path.basename(file_path1)
                    label_file_path1.config(text=file_name1)
                    print(file_path1)
                def select_path2():
                    global file_path2
                    file_path2 = filedialog.askopenfilename(
                        initialdir='.', title="Select A File", filetypes=(("Excel Files", "*.xlsx"), ("Excel Files", "*.xls")))
                    file_name2 = os.path.basename(file_path2)
                    label_file_path2.config(text=file_name2)
                    print(file_path2)
                def select_path3():
                    global file_path3
                    file_path3 = filedialog.askopenfilename(
                        initialdir='.', title="Select A File", filetypes=(("Excel Files", "*.xlsx"), ("Excel Files", "*.xls")))
                    file_name3 = os.path.basename(file_path3)
                    label_file_path3.config(text=file_name3)
                    print(file_path3)
                def select_path4():
                    global file_path4
                    file_path4 = filedialog.askopenfilename(
                        initialdir='.', title="Select A File", filetypes=(("Excel Files", "*.xlsx"), ("Excel Files", "*.xls")))
                    file_name4 = os.path.basename(file_path4)
                    label_file_path4.config(text=file_name4)
                    print(file_path4)
                def select_path5():
                    global file_path5
                    file_path5 = filedialog.askopenfilename(
                        initialdir='.', title="Select A File", filetypes=(("Excel Files", "*.xlsx"), ("Excel Files", "*.xls")))
                    file_name5 = os.path.basename(file_path5)
                    label_file_path5.config(text=file_name5)
                    print(file_path5)
                def select_path6():
                    global file_path6
                    file_path6 = filedialog.askopenfilename(
                        initialdir='.', title="Select A File", filetypes=(("Excel Files", "*.xlsx"), ("Excel Files", "*.xls")))
                    file_name6 = os.path.basename(file_path6)
                    label_file_path6.config(text=file_name6)
                    print(file_path6)
                def select_path7():
                    global file_path7
                    file_path7 = filedialog.askopenfilename(
                        initialdir='.', title="Select A File", filetypes=(("Excel Files", "*.xlsx"), ("Excel Files", "*.xls")))
                    file_name7 = os.path.basename(file_path7)
                    label_file_path7.config(text=file_name7)
                    print(file_path7)
                def select_path8():
                    global file_path8
                    file_path8 = filedialog.askopenfilename(
                        initialdir='.', title="Select A File", filetypes=(("Excel Files", "*.xlsx"), ("Excel Files", "*.xls")))
                    file_name8 = os.path.basename(file_path8)
                    label_file_path8.config(text=file_name8)
                    print(file_path8)
                def select_path9():
                    global file_path9
                    file_path9 = filedialog.askopenfilename(
                        initialdir='.', title="Select A File", filetypes=(("Excel Files", "*.xlsx"), ("Excel Files", "*.xls")))
                    file_name9 = os.path.basename(file_path9)
                    label_file_path9.config(text=file_name9)
                    print(file_path9)
                def select_path10():
                    global file_path10
                    file_path10 = filedialog.askopenfilename(
                        initialdir='.', title="Select A File", filetypes=(("Excel Files", "*.xlsx"), ("Excel Files", "*.xls")))
                    file_name10 = os.path.basename(file_path10)
                    label_file_path10.config(text=file_name10)
                    print(file_path10)
                    
                filters_select_path1 = tk.Button(filters_frame, text='Seleccionar archivo', command=select_path1)
                filters_select_path1.grid(row=0, column=2, padx=10, pady = '10')	
                filters_select_path2 = tk.Button(filters_frame, text='Seleccionar archivo', command=select_path2)
                filters_select_path2.grid(row=1, column=2, padx=10, pady = '10')
                filters_select_path3 = tk.Button(filters_frame, text='Seleccionar archivo', command=select_path3)
                filters_select_path3.grid(row=2, column=2, padx= 10, pady = '10')
                filters_select_path4 = tk.Button(filters_frame, text='Seleccionar archivo', command=select_path4)
                filters_select_path4.grid(row=3, column=2, padx= 10, pady = '10')
                filters_select_path5 = tk.Button(filters_frame, text='Seleccionar archivo', command=select_path5)
                filters_select_path5.grid(row=4, column=2, padx= 10, pady = '10')
                filters_select_path6 = tk.Button(filters_frame, text='Seleccionar archivo', command=select_path6)
                filters_select_path6.grid(row=5, column=2, padx= 10, pady = '10')
                filters_select_path7 = tk.Button(filters_frame, text='Seleccionar archivo', command=select_path7)
                filters_select_path7.grid(row=6, column=2, padx= 10, pady = '10')
                filters_select_path8 = tk.Button(filters_frame, text='Seleccionar archivo', command=select_path8)
                filters_select_path8.grid(row=7, column=2, padx= 10, pady = '10')
                filters_select_path9 = tk.Button(filters_frame, text='Seleccionar archivo', command=select_path9)
                filters_select_path9.grid(row=8, column=2, padx= 10, pady = '10')
                filters_select_path10 = tk.Button(filters_frame, text='Seleccionar archivo', command=select_path9)
                filters_select_path10.grid(row=9, column=2, padx= 10, pady = '10')
                filters_select_path1.config(font=('Cascadia Code', 9),fg = 'white', bg = 'black')
                filters_select_path2.config(font=('Cascadia Code', 9),fg = 'white', bg = 'black')
                filters_select_path3.config(font=('Cascadia Code', 9),fg = 'white', bg = 'black')
                filters_select_path4.config(font=('Cascadia Code', 9),fg = 'white', bg = 'black')
                filters_select_path5.config(font=('Cascadia Code', 9),fg = 'white', bg = 'black')
                filters_select_path6.config(font=('Cascadia Code', 9),fg = 'white', bg = 'black')
                filters_select_path7.config(font=('Cascadia Code', 9),fg = 'white', bg = 'black')
                filters_select_path8.config(font=('Cascadia Code', 9),fg = 'white', bg = 'black')
                filters_select_path9.config(font=('Cascadia Code', 9),fg = 'white', bg = 'black')
                filters_select_path10.config(font=('Cascadia Code', 9),fg = 'white', bg = 'black')

                label_file_path1 = tk.Label(filters_frame, text='Aun no se ha seleccionado ningun archivo')
                label_file_path1.grid(row=0, column=3, padx=10)
                label_file_path1.config(font=('Cascadia Code', 9),fg = 'black')
                label_file_path2 = tk.Label(filters_frame, text='Aun no se ha seleccionado ningun archivo')
                label_file_path2.grid(row=1, column=3, padx=10)
                label_file_path2.config(font=('Cascadia Code', 9),fg = 'black')
                label_file_path3 = tk.Label(filters_frame, text='Aun no se ha seleccionado ningun archivo')
                label_file_path3.grid(row=2, column=3, padx=10)
                label_file_path3.config(font=('Cascadia Code', 9),fg = 'black')
                label_file_path4 = tk.Label(filters_frame, text='Aun no se ha seleccionado ningun archivo')
                label_file_path4.grid(row=3, column=3, padx=10)
                label_file_path4.config(font=('Cascadia Code', 9),fg = 'black')
                label_file_path5 = tk.Label(filters_frame, text='Aun no se ha seleccionado ningun archivo')
                label_file_path5.grid(row=4, column=3, padx=10)
                label_file_path5.config(font=('Cascadia Code', 9),fg = 'black')
                label_file_path6 = tk.Label(filters_frame, text='Aun no se ha seleccionado ningun archivo')
                label_file_path6.grid(row=5, column=3, padx=10)
                label_file_path6.config(font=('Cascadia Code', 9),fg = 'black')
                label_file_path7 = tk.Label(filters_frame, text='Aun no se ha seleccionado ningun archivo')
                label_file_path7.grid(row=6, column=3, padx=10)
                label_file_path7.config(font=('Cascadia Code', 9),fg = 'black')
                label_file_path8 = tk.Label(filters_frame, text='Aun no se ha seleccionado ningun archivo')
                label_file_path8.grid(row=7, column=3, padx=10)
                label_file_path8.config(font=('Cascadia Code', 9),fg = 'black')
                label_file_path9 = tk.Label(filters_frame, text='Aun no se ha seleccionado ningun archivo')
                label_file_path9.grid(row=8, column=3, padx=10)
                label_file_path9.config(font=('Cascadia Code',9),fg = 'black')
                label_file_path10 = tk.Label(filters_frame, text='Aun no se ha seleccionado ningun archivo')
                label_file_path10.grid(row=9, column=3, padx=10)
                label_file_path10.config(font=('Cascadia Code',9),fg = 'black')
                label_select_device_system1 = tk.Label(filters_frame, text='iOS o Android?')  
                label_select_device_system1.grid(row=0, column=4, padx=10)
                label_select_device_system1.config(font=('Cascadia Code', 9),fg = 'black')
                label_select_device_system2 = tk.Label(filters_frame, text='iOS o Android?')
                label_select_device_system2.grid(row=1, column=4, padx=10)
                label_select_device_system2.config(font=('Cascadia Code', 9),fg = 'black')
                label_select_device_system3 = tk.Label(filters_frame, text='iOS o Android?')
                label_select_device_system3.grid(row=2, column=4, padx=10)
                label_select_device_system3.config(font=('Cascadia Code', 9),fg = 'black')
                label_select_device_system4 = tk.Label(filters_frame, text='iOS o Android?')
                label_select_device_system4.grid(row=3, column=4, padx=10)
                label_select_device_system4.config(font=('Cascadia Code', 9),fg = 'black')
                label_select_device_system5 = tk.Label(filters_frame, text='iOS o Android?')
                label_select_device_system5.grid(row=4, column=4, padx=10)
                label_select_device_system5.config(font=('Cascadia Code', 9),fg = 'black')
                label_select_device_system6 = tk.Label(filters_frame, text='iOS o Android?')
                label_select_device_system6.grid(row=5, column=4, padx=10)
                label_select_device_system6.config(font=('Cascadia Code', 9),fg = 'black')
                label_select_device_system7 = tk.Label(filters_frame, text='iOS o Android?')
                label_select_device_system7.grid(row=6, column=4, padx=10)
                label_select_device_system7.config(font=('Cascadia Code', 9),fg = 'black')
                label_select_device_system8 = tk.Label(filters_frame, text='iOS o Android?')
                label_select_device_system8.grid(row=7, column=4, padx=10)
                label_select_device_system8.config(font=('Cascadia Code', 9),fg = 'black')
                label_select_device_system9 = tk.Label(filters_frame, text='iOS o Android?')
                label_select_device_system9.grid(row=8, column=4, padx=10)
                label_select_device_system9.config(font=('Cascadia Code', 9),fg = 'black')
                label_select_device_system10 = tk.Label(filters_frame, text='iOS o Android?')
                label_select_device_system10.grid(row=9, column=4, padx=10)
                label_select_device_system10.config(font=('Cascadia Code', 9),fg = 'black')
                select_device_system1 = ttk.Combobox(filters_frame, width=10, state='readonly', values=['iOS', 'Android'])
                select_device_system1.grid(row=0, column=5, padx=10)
                select_device_system2 = ttk.Combobox(filters_frame, width=10, state='readonly', values=['iOS', 'Android'])
                select_device_system2.grid(row=1, column=5, padx=10)
                select_device_system3 = ttk.Combobox(filters_frame, width=10, state='readonly', values=['iOS', 'Android'])
                select_device_system3.grid(row=2, column=5, padx=10)
                select_device_system4 = ttk.Combobox(filters_frame, width=10, state='readonly', values=['iOS', 'Android'])
                select_device_system4.grid(row=3, column=5, padx=10)
                select_device_system5 = ttk.Combobox(filters_frame, width=10, state='readonly', values=['iOS', 'Android'])
                select_device_system5.grid(row=4, column=5, padx=10)
                select_device_system6 = ttk.Combobox(filters_frame, width=10, state='readonly', values=['iOS', 'Android'])
                select_device_system6.grid(row=5, column=5, padx=10)
                select_device_system7 = ttk.Combobox(filters_frame, width=10, state='readonly', values=['iOS', 'Android'])
                select_device_system7.grid(row=6, column=5, padx=10)
                select_device_system8 = ttk.Combobox(filters_frame, width=10, state='readonly', values=['iOS', 'Android'])
                select_device_system8.grid(row=7, column=5, padx=10)
                select_device_system9 = ttk.Combobox(filters_frame, width=10, state='readonly', values=['iOS', 'Android'])
                select_device_system9.grid(row=8, column=5, padx=10)
                select_device_system10 = ttk.Combobox(filters_frame, width=10, state='readonly', values=['iOS', 'Android'])
                select_device_system10.grid(row=9, column=5, padx=10)
                button_add_filter1 = tk.Button(filters_frame, text='Agregar filtro', command= lambda:  add_filter(path=f'{file_path1}', system= select_device_system1.get(), sensibility= sensibility.get(), phone=1))
                button_add_filter1.grid(row=0, column=6, padx=10)
                button_add_filter1.config(font=('Cascadia Code', 9),fg = 'white',bg ='green') 
                button_add_filter2 = tk.Button(filters_frame, text='Agregar filtro', command= lambda: add_filter(path=file_path2, system= select_device_system2.get() , sensibility=sensibility.get(), phone=2))
                button_add_filter2.grid(row=1, column=6, padx=10)
                button_add_filter2.config(font=('Cascadia Code', 9),fg = 'white',bg ='green')
                button_add_filter3 = tk.Button(filters_frame, text='Agregar filtro', command= lambda: add_filter(path=file_path3, system=select_device_system3.get() , sensibility=sensibility.get(), phone=3))
                button_add_filter3.grid(row=2, column=6, padx=10)
                button_add_filter3.config(font=('Cascadia Code', 9),fg = 'white',bg ='green')
                button_add_filter4 = tk.Button(filters_frame, text='Agregar filtro', command=lambda:  add_filter(path=file_path4, system=select_device_system4.get() , sensibility=sensibility.get(), phone=4))
                button_add_filter4.grid(row=3, column=6, padx=10)
                button_add_filter4.config(font=('Cascadia Code', 9),fg = 'white',bg ='green')
                button_add_filter5 = tk.Button(filters_frame, text='Agregar filtro', command= lambda: add_filter(path=file_path5, system= select_device_system5.get(), sensibility=sensibility.get(), phone=5))
                button_add_filter5.grid(row=4, column=6, padx=10)
                button_add_filter5.config(font=('Cascadia Code', 9),fg = 'white',bg ='green')
                button_add_filter6 = tk.Button(filters_frame, text='Agregar filtro', command=lambda:  add_filter(path=file_path6, system=select_device_system6.get() , sensibility=sensibility.get(), phone=6))
                button_add_filter6.grid(row=5, column=6, padx=10)
                button_add_filter6.config(font=('Cascadia Code', 9),fg = 'white',bg ='green')
                button_add_filter7 = tk.Button(filters_frame, text='Agregar filtro', command= lambda: add_filter(path=file_path7, system=select_device_system7.get() , sensibility=sensibility.get(), phone=7))
                button_add_filter7.grid(row=6, column=6, padx=10)
                button_add_filter7.config(font=('Cascadia Code', 9),fg = 'white',bg ='green')
                button_add_filter8 = tk.Button(filters_frame, text='Agregar filtro', command=lambda: add_filter(path=file_path8, system=select_device_system8.get() , sensibility=sensibility.get(), phone=8))
                button_add_filter8.grid(row=7, column=6, padx=10)
                button_add_filter8.config(font=('Cascadia Code', 9),fg = 'white',bg ='green')
                button_add_filter9 = tk.Button(filters_frame, text='Agregar filtro', command= lambda:add_filter(path=file_path9, system= select_device_system9.get(), sensibility=sensibility.get(), phone=9))
                button_add_filter9.grid(row=8, column=6, padx=10)
                button_add_filter9.config(font=('Cascadia Code', 9),fg = 'white',bg ='green')
                button_add_filter10 = tk.Button(filters_frame, text='Agregar filtro', command= lambda:add_filter(path=file_path10, system= select_device_system9.get(), sensibility=sensibility.get(), phone=10))
                button_add_filter10.grid(row=9, column=6, padx=10)
                button_add_filter10.config(font=('Cascadia Code', 9),fg = 'white',bg ='green')


  
            def wave_arrival_time():
                try: 
                    err = main_ws['A2'].value
                    if  err != None: 
                       main_ws['M1'] = 'Distance between phones (m)'
                       main_ws['L1'] = 'Wave arrival time (s)'
                       increment= 0
                       for i in range (main_ws.max_row -2):
                               main_ws['L2'] = 0 
                               main_ws['M2'] = 0
                               k_b = 'K' + str(i+2)
                               k_a = 'K' + str(i+3)
                               main_ws['L' + str(i+3)] = f'=ABS({k_b}-{k_a})'
                               increment = increment + 5
                               main_ws['M' + str(i+3)] = increment
                               #make m column with distance between phones 5m per phone
                       main_wb.save(output_path)
                       #create a xy chart with openpyxl
                       chart = ScatterChart()
                       chart.title = 'Wave arrival time'
                       chart.style = 13
                       chart.x_axis.title = 'Distancia entre celulares (m)'
                       chart.y_axis.title = 'Tiempo de llegada (s)'
                       xvalues = Reference(main_ws, min_col=13, min_row=2, max_row=main_ws.max_row)
                       values = Reference(main_ws, min_col=12, min_row=1, max_row=main_ws.max_row)
                       series = Series(values, xvalues, title_from_data=True)
                       chart.series.append(series)
                       main_ws.add_chart(chart, "O2")
                       main_wb.save(output_path)
                       main_wb.close()
                       os.startfile(output_path)
                       print(err)
                    else:
                          messagebox.showwarning('Aviso', 'Parece que no has filtrado ningún archivo, obtendras un excel sin datos')
                          main_ws['M1'] = 'Distance between phones (m)'
                          main_ws['L1'] = 'Wave arrival time (s)'
                          increment= 0
                          for i in range (main_ws.max_row -2):
                                  main_ws['L2'] = 0 
                                  main_ws['M2'] = 0
                                  k_b = 'K' + str(i+2)
                                  k_a = 'K' + str(i+3)
                                  main_ws['L' + str(i+3)] = f'=ABS({k_b}-{k_a})'
                                  increment = increment + 5
                                  main_ws['M' + str(i+3)] = increment
                                  
                          main_wb.save(output_path)
                          main_wb.close()
                          os.startfile(output_path)
                          print(err)

                except:
                    messagebox.showerror('Error', 'Primero tienes que crear el archivo de salida')

                
            
            def return_function():
                    phyphox_sismic_exploration_filter.destroy()
                    main()
                

            header = tk.Frame( phyphox_sismic_exploration_filter)
            header.pack()
            # aumentar el tamaño de la fuente
            label_title = tk.Label(
                header, text="Phyphox Sismic Exploration Filter", font=("Cascadia Code", 20))
            label_title.grid(row=0, column=0, pady=10)
            label_title.config(font=("Cascadia Code", 20))
            label_description = tk.Label(
                header, text="Este programa almacena los datos filtrados usando la logica de Phyphox anomaly detector en un archivo de excel.", font=("Cascadia code", 12))
            label_description.grid(row=1, column=0, padx=0, pady=10, sticky='NSEW')
            sensibility_frame = tk.Frame(phyphox_sismic_exploration_filter)
            sensibility_frame.pack()
            sensibility = tk.DoubleVar()
            sensibility.set(0.5)
            label_sensibility = tk.Label(
               sensibility_frame, text='Ajusta la sensibilidad del filtro', font=("Cascadia Code", 10))
            label_sensibility.grid(row=4, column=0)
        
            label_sensibility_value = tk.Label(
                sensibility_frame, textvariable=sensibility, font=("Cascadia Code", 10))
            label_sensibility_value.grid(row=5, column=0)
        
            slider_sensibility = ttk.Scale(
               sensibility_frame, from_=0, to=1, variable=sensibility, orient=tk.HORIZONTAL, length=200)
            slider_sensibility.grid(row=6, column=0)  
            #     ke sensibility a three decimal numbe
    
    
            label_sys = tk.Label(header, text="Ingresa el sistema operativo del celular, puede ser iOS o Android", font=(
                "Cascadia code", 10))
            label_sys.grid(row=2, column=0, pady=5)
            buttons = tk.Frame(phyphox_sismic_exploration_filter)
            buttons.pack()
            global button_Start_Show_Filter
            button_Start_Show_Filter = tk.Button(
                buttons, text='Empezar', command=Start_filter)
            button_Start_Show_Filter.grid(row=0, column=3, padx=10, pady=10)
            button_Start_Show_Filter.config(font=("Cascadia Code", 10),
                                    width=20, height=1, bg="red", fg="white")
            output_name = tk.Label(buttons, text='Nombre del archivo de salida')
            output_name.grid(row=0, column=4, padx=10, pady=10)
            output_name.config(font=("Cascadia Code", 10),
                               width=35, height=1, bg="black", fg="white")
            global output_name_entry
            output_name_entry = tk.Entry(
                buttons, width=20, font=("Cascadia Code", 10))
            output_name_entry.grid(row=0, column=5, padx=10, pady=10)
            output_name_entry.config(font=("Cascadia Code", 10))

            create_excel_button = tk.Button(
                buttons, text='Crear Excel', command= lambda : create_excel(output_name_entry.get()))
            create_excel_button.grid(row=0, column=6, padx=10, pady=10)
            create_excel_button.config(font=("Cascadia Code", 10),
                                       width=20, height=1, bg="green", fg="white")
            
            save_excel_button = tk.Button( buttons, text='Guardar Excel', command=save_file)
            save_excel_button.grid(row=0, column=7, padx=10, pady=10)
            save_excel_button.config(font=("Cascadia Code", 10),
                                       width=20, height=1, bg="blue", fg="white")
            finalizar_button = tk.Button( buttons, text='Finalizar filtro', command= wave_arrival_time)
            finalizar_button.grid(row=0, column=8, padx=10, pady=10)
            finalizar_button.config(font=("Cascadia Code", 10),
                                       width=20, height=1, bg="black", fg="white")
    
    
            

            exit_button = tk.Button(phyphox_sismic_exploration_filter, text='Regresar', command= return_function)
            exit_button.pack(side='bottom', pady=10, anchor='s')
            exit_button.config(font=("Cascadia Code", 10), width=20,
                               bg="black", fg="white") 

                
    def return_function_root():
        
        phyphox_sismic_exploration_main.destroy()
        import phyphox_connection as phy
        phy.Phyphox_connection.mainloop()

    # create tkinter window
    phyphox_sismic_exploration_main = ThemedTk(theme='arc')
    phyphox_sismic_exploration_main.title("Phyphox anomaly detector")
    #remove icon 

    # make the window full screen
    phyphox_sismic_exploration_main.state('zoomed')
    phyphox_sismic_exploration_main.attributes('-fullscreen', True)
    phyphox_sismic_exploration_main.resizable(False, False)
    phyphox_sismic_exploration_main.config(padx=20, pady=20)
    # define useful variable
    global file_path
    file_path = tk.StringVar(value='')

    # create a frame
    header = tk.Frame(phyphox_sismic_exploration_main)
    header.pack()
    # aumentar el tamaño de la fuente

    label_title = tk.Label(
        header, text="Phyphox anomaly detector", font=("Cascadia Code", 20))
    label_title.grid(row=0, column=0, pady=30, sticky='n')
    label_title.config(font=("Cascadia Code", 20))
    label_description = tk.Label(
        header, text="Este programa filtra el momento exacto en el que el acelerómetro ha detectado una anomalía.", font=("Cascadia code", 12))
    label_description.grid(row=1, column=0, padx=0, pady=20, sticky='NSEW')
    label_sys = tk.Label(header, text="Ingresa el sistema operativo del celular, puede ser iOS o Android", font=(
        "Cascadia code", 10))
    label_sys.grid(row=2, column=0, pady=5)
    # create a option menu
    entry = ttk.Combobox(header, values=['iOS', 'Android'], state='readonly', width=20, height=2, font=('Cascadia Code', 12), justify='center')
    entry.grid(row=3, column=0, pady=5)
 
    # create a button


    # create a slider for the sensibility and show the value
    sensibility = tk.DoubleVar()
    sensibility.set(0.5)
    # make sensibility a three decimal numbe

    label_sensibility = tk.Label(
        header, text='Ajusta la sensibilidad del filtro', font=("Cascadia Code", 10))
    label_sensibility.grid(row=4, column=0, pady=5)

    label_sensibility_value = tk.Label(
        header, textvariable=sensibility, font=("Cascadia Code", 10))
    label_sensibility_value.grid(row=5, column=0, pady=5)

    slider_sensibility = ttk.Scale(
        header, from_=0, to=1, variable=sensibility, orient=tk.HORIZONTAL, length=200)
    slider_sensibility.grid(row=6, column=0, pady=5)

    AcUn_buttons = tk.Frame(phyphox_sismic_exploration_main)
    # mover los botones mas abajo
    AcUn_buttons.pack(pady=10)
    # create buttons

    ManageData_buttons = tk.Frame(phyphox_sismic_exploration_main)
    ManageData_buttons.pack()
    # Buttons

    button_select = tk.Button(
        ManageData_buttons, text="Seleccionar archivo de excel", command=select_file)
    button_select.grid(row=3, column=0, pady=10)
    button_read = tk.Button(
        ManageData_buttons, text="Leer archivo", command=read_file)
    button_read.grid(row=3, column=1, padx=10, pady=10)
    button_filter = tk.Button(
        ManageData_buttons, text="Filtrar", command=filter_data)
    button_filter.grid(row=3, column=2, padx=10, pady=10)
    # create button to add data to the excel file

    # styling  the buttons

    button_select.config(font=("Cascadia Code", 10),
                         width=30, height=1, bg="blue", fg="white")
    button_read.config(font=("Cascadia Code", 10),
                       width=15, height=1, bg="blue", fg="white")
    button_filter.config(font=("Cascadia Code", 10),
                         width=10, height=1, bg="orange", fg="white")

    # create a separator
    separator = ttk.Separator(phyphox_sismic_exploration_main, orient='horizontal')
    separator.pack(fill='x', pady=10)
    # create a output frame

    output = tk.Frame(phyphox_sismic_exploration_main)
    output.pack()
 
    
    label_file_path = tk.Label(output, text='Salida')
    label_file_path.grid(row=0, column=0, padx=10, pady=10, sticky='n')
    label_file_path.config(font=('Cascadia Code', 10, 'bold'),
                           bg='white', fg='black', relief='sunken', borderwidth=2)
    bottom = tk.Frame(phyphox_sismic_exploration_main)
    bottom.pack(side='bottom', pady=10, anchor='se')
    #exit button
    button_exit = tk.Button(bottom, text="Regresar", command=return_function_root)
    button_exit.grid(row=0, column=0, sticky='sw')
    button_exit.config(font=("Cascadia Code", 10),
                       width=45, height=1, bg="black", fg="white")
    

    #button for open the phyphox sismic exploration filter

    button_showfilter_menu = tk.Button(
        bottom, text='--> Phyphox Sismic Exploration Filter <--', command= (Show_add_filter))
    button_showfilter_menu.grid(row=2, column=0, pady=5)
    button_showfilter_menu.config(font=("Cascadia Code", 10),
                                  width=45, height=1, bg="red", fg="white")
    #create a credits label

    phyphox_sismic_exploration_main.mainloop()


if __name__ == '__main__':
    main()
